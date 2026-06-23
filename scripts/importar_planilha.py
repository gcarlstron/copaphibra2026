"""Importador inicial da planilha Copa Phibra 2026.

Lê o arquivo .xlsx e popula o banco com rodadas, jogos, usuários e palpites
da fase de grupos (linhas 2–73 da aba OFICIAL). Mata-mata (linha 74+) é
ignorado nesta versão.

Uso:
    python scripts/importar_planilha.py [SENHA_PADRAO]

    SENHA_PADRAO  senha provisória para os 10 jogadores (padrão: copaphibra2026).

Idempotente: re-executar não duplica registros — get-or-create em tudo.
Jogos já ENCERRADOS (e suas pontuações) são preservados, não sobrescritos — o
resultado no banco é autoritativo (pode ter vindo da ESPN/admin). A senha de
usuários já existentes também é preservada. O estado de palpite das rodadas
(aberta/janela) nunca é alterado no re-import. O usuário admin pré-existente
(criado por scripts/criar_admin.py) não é tocado.

Requisitos:
    pip install openpyxl
"""

from __future__ import annotations

import sys
from dataclasses import dataclass
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Callable, NamedTuple

# ---------------------------------------------------------------------------
# Bootstrap de sys.path para permitir ``import app...`` ao rodar diretamente.
# ---------------------------------------------------------------------------
_PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_PROJECT_ROOT))

# Após o path bootstrap, as importações do app ficam disponíveis.
import openpyxl  # noqa: E402
from openpyxl.worksheet.worksheet import Worksheet  # noqa: E402
from sqlalchemy import select  # noqa: E402
from sqlalchemy.orm import Session  # noqa: E402

from app.database import SessionLocal  # noqa: E402
from app.models import Jogo, Palpite, Rodada, Usuario  # noqa: E402
from app.services.auth import hash_senha  # noqa: E402
from app.services.dashboard import STATUS_AGENDADO, STATUS_ENCERRADO  # noqa: E402
from app.services.scoring import calcular_pontos  # noqa: E402

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

SENHA_PADRAO: str = "copaphibra2026"

XLSX_PATH: Path = _PROJECT_ROOT / "import" / "COPA PHIBRA 2026 OFICIAL ATÉ A FINAL TERCEIRA FASE.xlsx"

# Aba com resultados oficiais e placar
ABA_OFICIAL: str = "OFICIAL"

# 10 abas de jogadores ativos; key = nome da aba, value = nome de exibição
JOGADORES: dict[str, str] = {
    "BERNARDO": "Bernardo",
    "FERNANDO": "Fernando",
    "GABRIEL": "Gabriel",
    "GUSTAVO": "Gustavo",
    "MARCIO": "Marcio",
    "MARQUES": "Marques",
    "RENAN": "Renan",
    "RICARDO": "Ricardo",
    "THIAGO": "Thiago",
    "SOARES": "Soares",
}

# Limites de linhas na planilha (1-based, inclusive)
LINHA_INICIO: int = 2
LINHA_FIM: int = 73  # última linha da fase de grupos

# Definição das 3 rodadas da fase de grupos
RODADAS: list[tuple[int, str, int, int]] = [
    # (ordem, nome, linha_inicio, linha_fim)
    (1, "1ª Rodada", 2, 25),
    (2, "2ª Rodada", 26, 49),
    (3, "3ª Rodada", 50, 73),
]

# Totais esperados (R1+R2) para a validação final. Atualizados em 2026-06-23 a
# partir da col O da planilha da 3ª fase (e conferem com a classificação atual do
# banco). A R3 ainda não tem resultado, então não soma pontos aqui.
# Obs.: este snapshot precisa ser reajustado quando a R3 começar a ter resultados.
ESPERADO: dict[str, int] = {
    "Bernardo": 111,
    "Thiago": 118,
    "Ricardo": 101,
    "Fernando": 96,
    "Gustavo": 101,
    "Marcio": 120,
    "Gabriel": 85,
    "Renan": 83,
    "Soares": 119,
    "Marques": 129,
}


# ---------------------------------------------------------------------------
# Tipos auxiliares
# ---------------------------------------------------------------------------


class LinhaOficial(NamedTuple):
    """Dados de um jogo lidos da aba OFICIAL."""

    row: int
    data_hora: datetime
    time_casa: str
    time_visitante: str
    gols_casa: int | None
    gols_visitante: int | None
    status: str


class LinhaPalpite(NamedTuple):
    """Palpite de um jogador em uma linha."""

    row: int
    gols_casa: int
    gols_visitante: int
    pontos_planilha: int | None  # col O — só para log/diagnóstico


@dataclass(slots=True)
class LinhaValidacao:
    """Uma linha da validação de totais (por jogador)."""

    nome: str
    calculado: int
    esperado: int | None
    ok: bool


@dataclass(slots=True)
class ResultadoImportacao:
    """Resumo do que a importação fez — retornado por `importar()`.

    Não imprime nem aborta: o caller (CLI ou rota admin) decide o que mostrar e
    se trata `todos_ok=False` como erro.
    """

    arquivo: str
    usuarios_criados: int
    usuarios_atualizados: int
    rodadas_criadas: int
    rodadas_atualizadas: int
    jogos_criados: int
    jogos_atualizados: int
    jogos_protegidos: int
    palpites_criados: int
    palpites_atualizados: int
    divergencias: list[tuple]
    validacao: list[LinhaValidacao]
    todos_ok: bool


# ---------------------------------------------------------------------------
# Leitura da planilha
# ---------------------------------------------------------------------------


def _combinar_data_hora(data_cell: object, hora_cell: object) -> datetime:
    """Combina as colunas A (date) e B (time) em datetime UTC-aware."""
    if isinstance(data_cell, datetime):
        d = data_cell.date()
    elif isinstance(data_cell, date):
        d = data_cell
    else:
        raise ValueError(f"Coluna A inesperada: {data_cell!r}")

    if isinstance(hora_cell, time):
        h = hora_cell
    elif hora_cell is None:
        h = time(0, 0)
    else:
        raise ValueError(f"Coluna B inesperada: {hora_cell!r}")

    return datetime(d.year, d.month, d.day, h.hour, h.minute, tzinfo=timezone.utc)


def ler_oficial(ws: Worksheet) -> list[LinhaOficial]:
    """Lê as linhas 2–73 da aba OFICIAL e retorna LinhaOficial por linha."""
    linhas: list[LinhaOficial] = []
    for r in range(LINHA_INICIO, LINHA_FIM + 1):
        row = list(ws.iter_rows(min_row=r, max_row=r, min_col=1, max_col=7, values_only=True))[0]
        # Colunas: A=0 B=1 C=2 D=3 E=4 F=5 G=6
        data_cell, hora_cell = row[0], row[1]
        time_casa: str = str(row[2]) if row[2] is not None else ""
        time_visitante: str = str(row[6]) if row[6] is not None else ""

        data_hora = _combinar_data_hora(data_cell, hora_cell)

        gols_casa_raw = row[3]
        gols_vis_raw = row[5]

        if gols_casa_raw is not None and gols_vis_raw is not None:
            gols_casa: int | None = int(gols_casa_raw)
            gols_vis: int | None = int(gols_vis_raw)
            status = STATUS_ENCERRADO
        else:
            gols_casa = None
            gols_vis = None
            status = STATUS_AGENDADO

        linhas.append(
            LinhaOficial(
                row=r,
                data_hora=data_hora,
                time_casa=time_casa,
                time_visitante=time_visitante,
                gols_casa=gols_casa,
                gols_visitante=gols_vis,
                status=status,
            )
        )
    return linhas


def ler_palpites_jogador(
    ws: Worksheet,
    nome_aba: str,
    linhas_oficial: list[LinhaOficial],
) -> dict[int, LinhaPalpite]:
    """Lê os palpites de um jogador (linhas 2–73) após validar alinhamento.

    Retorna um dict row_number -> LinhaPalpite apenas para linhas onde o
    jogador tem AMBOS gols_casa e gols_visitante preenchidos.

    Levanta RuntimeError se a coluna C ou G não bater com a aba OFICIAL.
    """
    palpites: dict[int, LinhaPalpite] = {}
    for linha_of in linhas_oficial:
        r = linha_of.row
        row = list(ws.iter_rows(min_row=r, max_row=r, min_col=1, max_col=15, values_only=True))[0]
        # Validação de alinhamento: col C e G devem bater com OFICIAL
        tc_jogador = str(row[2]) if row[2] is not None else ""
        tv_jogador = str(row[6]) if row[6] is not None else ""
        if tc_jogador != linha_of.time_casa or tv_jogador != linha_of.time_visitante:
            raise RuntimeError(
                f"Desalinhamento na aba {nome_aba!r} linha {r}: "
                f"esperado {linha_of.time_casa!r} x {linha_of.time_visitante!r}, "
                f"encontrado {tc_jogador!r} x {tv_jogador!r}. "
                "Abortar — não importe palpites desalinhados."
            )
        d_raw, f_raw = row[3], row[5]
        if d_raw is None or f_raw is None:
            continue  # jogador não palpitou neste jogo
        palpites[r] = LinhaPalpite(
            row=r,
            gols_casa=int(d_raw),
            gols_visitante=int(f_raw),
            pontos_planilha=row[14],  # col O — só para log
        )
    return palpites


# ---------------------------------------------------------------------------
# Operações de banco (get-or-create)
# ---------------------------------------------------------------------------


def _get_or_create_usuario(
    db: Session,
    username: str,
    nome: str,
    senha: str,
) -> tuple[Usuario, bool]:
    """Retorna (usuario, criado).

    Usuário NOVO: cria com a senha provisória.
    Usuário EXISTENTE: atualiza nome e reativa, mas **NUNCA** reescreve
    `senha_hash` — preserva a senha que o jogador já trocou (idempotência de
    credencial) — nem toca em `is_admin`.
    """
    stmt = select(Usuario).where(Usuario.username == username)
    usuario = db.execute(stmt).scalar_one_or_none()
    if usuario is None:
        usuario = Usuario(
            nome=nome,
            username=username,
            senha_hash=hash_senha(senha),
            is_admin=False,
            ativo=True,
        )
        db.add(usuario)
        return usuario, True
    # Já existe: atualiza nome e reativa, mas NUNCA reescreve senha_hash
    # (preserva a senha que o jogador já trocou) nem toca em is_admin.
    usuario.nome = nome
    usuario.ativo = True
    return usuario, False


def _get_or_create_rodada(
    db: Session,
    ordem: int,
    nome: str,
) -> tuple[Rodada, bool]:
    """Retorna (rodada, criado). Atualiza nome se já existir."""
    stmt = select(Rodada).where(Rodada.ordem == ordem)
    rodada = db.execute(stmt).scalar_one_or_none()
    if rodada is None:
        rodada = Rodada(
            nome=nome,
            ordem=ordem,
            aberta=False,
            abertura=None,
            fechamento=None,
        )
        db.add(rodada)
        return rodada, True
    rodada.nome = nome
    return rodada, False


def _get_or_create_jogo(
    db: Session,
    rodada_id: int,
    time_casa: str,
    time_visitante: str,
    data_hora: datetime,
    gols_casa: int | None,
    gols_visitante: int | None,
    status: str,
) -> tuple[Jogo, bool, bool]:
    """Retorna (jogo, criado, protegido).

    - Jogo NOVO: cria com os dados da planilha → (jogo, True, False).
    - Jogo já ENCERRADO: NÃO sobrescreve placar/status/data — o resultado no
      banco é autoritativo (pode ter vindo da ESPN/admin) → (jogo, False, True).
    - Jogo existente ainda não encerrado: atualiza data_hora/gols/status.
    """
    stmt = select(Jogo).where(
        Jogo.rodada_id == rodada_id,
        Jogo.time_casa == time_casa,
        Jogo.time_visitante == time_visitante,
    )
    jogo = db.execute(stmt).scalar_one_or_none()
    if jogo is None:
        jogo = Jogo(
            rodada_id=rodada_id,
            data_hora=data_hora,
            time_casa=time_casa,
            time_visitante=time_visitante,
            gols_casa=gols_casa,
            gols_visitante=gols_visitante,
            status=status,
        )
        db.add(jogo)
        return jogo, True, False
    if jogo.status == STATUS_ENCERRADO:
        # Não sobrescreve um jogo já encerrado (resultado autoritativo no banco).
        return jogo, False, True
    jogo.data_hora = data_hora
    jogo.gols_casa = gols_casa
    jogo.gols_visitante = gols_visitante
    jogo.status = status
    return jogo, False, False


def _get_or_create_palpite(
    db: Session,
    usuario_id: int,
    jogo_id: int,
    gols_casa: int,
    gols_visitante: int,
    pontos: int,
) -> tuple[Palpite, bool]:
    """Retorna (palpite, criado). Atualiza gols e pontos se já existir."""
    stmt = select(Palpite).where(
        Palpite.usuario_id == usuario_id,
        Palpite.jogo_id == jogo_id,
    )
    palpite = db.execute(stmt).scalar_one_or_none()
    if palpite is None:
        palpite = Palpite(
            usuario_id=usuario_id,
            jogo_id=jogo_id,
            gols_casa=gols_casa,
            gols_visitante=gols_visitante,
            pontos=pontos,
        )
        db.add(palpite)
        return palpite, True
    palpite.gols_casa = gols_casa
    palpite.gols_visitante = gols_visitante
    palpite.pontos = pontos
    return palpite, False


# ---------------------------------------------------------------------------
# Lógica principal
# ---------------------------------------------------------------------------


def importar(
    senha: str = SENHA_PADRAO,
    session_factory: Callable[[], Session] | None = None,
    xlsx_path: Path | None = None,
) -> ResultadoImportacao:
    """Executa a importação completa da planilha no banco e devolve um resumo.

    Idempotente. NÃO imprime nem chama ``sys.exit`` — devolve um
    ``ResultadoImportacao`` (com ``todos_ok``) e o caller decide o que fazer.
    Levanta ``FileNotFoundError``/``RuntimeError`` em erro de arquivo/aba, e
    ``RuntimeError`` em desalinhamento de palpite — seguro para chamar de uma
    rota web.

    Args:
        senha: Senha provisória aplicada a todos os 10 jogadores.
        session_factory: Fábrica de sessão SQLAlchemy. Padrão: SessionLocal
            do app (banco de produção). Passe um sessionmaker diferente para
            testes.
        xlsx_path: Caminho do arquivo .xlsx. Padrão: XLSX_PATH.
    """
    _xlsx = xlsx_path if xlsx_path is not None else XLSX_PATH
    _session_factory = session_factory if session_factory is not None else SessionLocal

    if not _xlsx.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {_xlsx}")

    print(f"Abrindo planilha: {_xlsx.name}")
    wb = openpyxl.load_workbook(str(_xlsx), data_only=True)

    # Valida que as abas esperadas existem
    for aba in [ABA_OFICIAL] + list(JOGADORES.keys()):
        if aba not in wb.sheetnames:
            raise RuntimeError(f"Aba {aba!r} não encontrada na planilha.")

    ws_oficial = wb[ABA_OFICIAL]

    # Lê dados da aba OFICIAL
    linhas_oficial = ler_oficial(ws_oficial)
    print(f"Lidas {len(linhas_oficial)} linhas da aba OFICIAL (linhas 2–73).")
    jogos_com_resultado = sum(1 for l in linhas_oficial if l.gols_casa is not None)
    jogos_sem_resultado = len(linhas_oficial) - jogos_com_resultado
    print(f"  {jogos_com_resultado} jogos com resultado, {jogos_sem_resultado} agendados.")

    # Lê palpites de cada jogador (com validação de alinhamento)
    palpites_por_jogador: dict[str, dict[int, LinhaPalpite]] = {}
    for aba_nome in JOGADORES:
        ws_jogador = wb[aba_nome]
        palpites = ler_palpites_jogador(ws_jogador, aba_nome, linhas_oficial)
        palpites_por_jogador[aba_nome] = palpites
        print(f"  {aba_nome}: {len(palpites)} palpites encontrados.")

    # Mapeia row -> LinhaOficial para lookup rápido
    linha_por_row: dict[int, LinhaOficial] = {l.row: l for l in linhas_oficial}

    # Monta mapa de rodada: row -> (ordem, nome)
    row_para_rodada: dict[int, tuple[int, str]] = {}
    for ordem, nome, r_ini, r_fim in RODADAS:
        for r in range(r_ini, r_fim + 1):
            row_para_rodada[r] = (ordem, nome)

    db = _session_factory()
    try:
        # ---------------------------------------------------------------
        # 1. Usuários
        # ---------------------------------------------------------------
        stats_usuarios = {"criados": 0, "atualizados": 0}
        usuario_por_aba: dict[str, Usuario] = {}

        for aba_nome, nome_display in JOGADORES.items():
            username = nome_display.lower()
            usuario, criado = _get_or_create_usuario(db, username, nome_display, senha)
            usuario_por_aba[aba_nome] = usuario
            if criado:
                stats_usuarios["criados"] += 1
            else:
                stats_usuarios["atualizados"] += 1

        db.flush()  # gera IDs antes de criar rodadas/jogos

        # ---------------------------------------------------------------
        # 2. Rodadas
        # ---------------------------------------------------------------
        stats_rodadas = {"criadas": 0, "atualizadas": 0}
        rodada_por_ordem: dict[int, Rodada] = {}

        for ordem, nome, _ri, _rf in RODADAS:
            rodada, criada = _get_or_create_rodada(db, ordem, nome)
            rodada_por_ordem[ordem] = rodada
            if criada:
                stats_rodadas["criadas"] += 1
            else:
                stats_rodadas["atualizadas"] += 1

        db.flush()  # gera IDs de rodadas antes de criar jogos

        # ---------------------------------------------------------------
        # 3. Jogos
        # ---------------------------------------------------------------
        stats_jogos = {"criados": 0, "atualizados": 0, "protegidos": 0}
        # Mapa (rodada_id, time_casa, time_visitante) -> Jogo
        jogo_por_chave: dict[tuple[int, str, str], Jogo] = {}
        # Jogos já encerrados que NÃO foram tocados — seus palpites/pontos
        # também não devem ser reescritos (resultado autoritativo no banco).
        jogos_protegidos: set[tuple[int, str, str]] = set()

        for linha in linhas_oficial:
            rodada_info = row_para_rodada.get(linha.row)
            if rodada_info is None:
                # Linha fora das 3 rodadas — não deveria ocorrer mas é seguro ignorar
                continue
            ordem, _ = rodada_info
            rodada = rodada_por_ordem[ordem]

            jogo, criado, protegido = _get_or_create_jogo(
                db=db,
                rodada_id=rodada.id,
                time_casa=linha.time_casa,
                time_visitante=linha.time_visitante,
                data_hora=linha.data_hora,
                gols_casa=linha.gols_casa,
                gols_visitante=linha.gols_visitante,
                status=linha.status,
            )
            chave = (rodada.id, linha.time_casa, linha.time_visitante)
            jogo_por_chave[chave] = jogo
            if protegido:
                jogos_protegidos.add(chave)
                stats_jogos["protegidos"] += 1
            elif criado:
                stats_jogos["criados"] += 1
            else:
                stats_jogos["atualizados"] += 1

        db.flush()  # gera IDs de jogos antes de criar palpites

        # ---------------------------------------------------------------
        # 4. Palpites
        # ---------------------------------------------------------------
        stats_palpites = {"criados": 0, "atualizados": 0}
        # Diagnóstico: linhas onde nossa pontuação difere da col O da planilha
        divergencias_planilha: list[tuple[str, int, int, int, int, int, int, int]] = []

        for aba_nome, palpites_aba in palpites_por_jogador.items():
            usuario = usuario_por_aba[aba_nome]
            for row_num, lp in palpites_aba.items():
                linha_of = linha_por_row[row_num]
                rodada_info = row_para_rodada[row_num]
                ordem, _ = rodada_info
                rodada = rodada_por_ordem[ordem]
                chave_jogo = (rodada.id, linha_of.time_casa, linha_of.time_visitante)
                jogo = jogo_por_chave.get(chave_jogo)
                if jogo is None:
                    # Não deveria ocorrer; jogo foi criado na etapa anterior
                    print(f"AVISO: jogo não encontrado para linha {row_num} aba {aba_nome}")
                    continue
                if chave_jogo in jogos_protegidos:
                    # Jogo encerrado preservado — não reescreve palpite/pontos
                    # (evita recomputar pontos a partir de um placar divergente).
                    continue

                # Calcula pontos
                if linha_of.gols_casa is not None and linha_of.gols_visitante is not None:
                    pontos = calcular_pontos(
                        lp.gols_casa,
                        lp.gols_visitante,
                        linha_of.gols_casa,
                        linha_of.gols_visitante,
                    )
                else:
                    pontos = 0

                # Diagnóstico vs col O (só para jogos com resultado)
                if (
                    lp.pontos_planilha is not None
                    and linha_of.gols_casa is not None
                    and pontos != lp.pontos_planilha
                ):
                    divergencias_planilha.append((
                        aba_nome,
                        row_num,
                        lp.gols_casa,
                        lp.gols_visitante,
                        linha_of.gols_casa,
                        linha_of.gols_visitante,
                        pontos,
                        lp.pontos_planilha,
                    ))

                _, criado = _get_or_create_palpite(
                    db=db,
                    usuario_id=usuario.id,
                    jogo_id=jogo.id,
                    gols_casa=lp.gols_casa,
                    gols_visitante=lp.gols_visitante,
                    pontos=pontos,
                )
                if criado:
                    stats_palpites["criados"] += 1
                else:
                    stats_palpites["atualizados"] += 1

        db.commit()

        # ---------------------------------------------------------------
        # 5. Validação de totais (R1+R2). NÃO imprime nem aborta — devolve o
        #    resultado; o caller (CLI ou rota admin) decide o que mostrar.
        # ---------------------------------------------------------------
        validacao: list[LinhaValidacao] = []
        todos_ok = True
        for aba_nome, nome_display in JOGADORES.items():
            usuario = usuario_por_aba[aba_nome]
            # Soma os pontos apenas dos palpites de jogos COM resultado.
            stmt_total = (
                select(Palpite.pontos)
                .join(Jogo, Palpite.jogo_id == Jogo.id)
                .where(
                    Palpite.usuario_id == usuario.id,
                    Jogo.status == STATUS_ENCERRADO,
                )
            )
            total_calc = sum(r[0] for r in db.execute(stmt_total).all())
            esperado = ESPERADO.get(nome_display)
            ok = esperado is not None and total_calc == esperado
            if esperado is not None and not ok:
                todos_ok = False
            validacao.append(LinhaValidacao(nome_display, total_calc, esperado, ok))

        return ResultadoImportacao(
            arquivo=_xlsx.name,
            usuarios_criados=stats_usuarios["criados"],
            usuarios_atualizados=stats_usuarios["atualizados"],
            rodadas_criadas=stats_rodadas["criadas"],
            rodadas_atualizadas=stats_rodadas["atualizadas"],
            jogos_criados=stats_jogos["criados"],
            jogos_atualizados=stats_jogos["atualizados"],
            jogos_protegidos=stats_jogos["protegidos"],
            palpites_criados=stats_palpites["criados"],
            palpites_atualizados=stats_palpites["atualizados"],
            divergencias=divergencias_planilha,
            validacao=validacao,
            todos_ok=todos_ok,
        )

    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def main() -> None:
    """Ponto de entrada do script (CLI): importa, imprime o resumo e sai 1 se divergir."""
    senha = sys.argv[1] if len(sys.argv) > 1 else SENHA_PADRAO
    res = importar(senha=senha)

    print(f"\n=== RESUMO DA IMPORTAÇÃO ({res.arquivo}) ===")
    print(f"Usuários : {res.usuarios_criados} criados, {res.usuarios_atualizados} atualizados")
    print(f"Rodadas  : {res.rodadas_criadas} criadas, {res.rodadas_atualizadas} atualizadas")
    print(
        f"Jogos    : {res.jogos_criados} criados, {res.jogos_atualizados} atualizados, "
        f"{res.jogos_protegidos} preservados (encerrados)"
    )
    print(f"Palpites : {res.palpites_criados} criados, {res.palpites_atualizados} atualizados")

    if res.divergencias:
        print(f"\nDIVERGÊNCIAS vs col O da planilha ({len(res.divergencias)} casos):")
        for aba, row, pc, pv, oc, ov, nosso, planO in res.divergencias:
            print(f"  {aba:12s} L{row}: palpite {pc}-{pv} | oficial {oc}-{ov} | nosso {nosso}, col O {planO}")
    else:
        print("\nNenhuma divergência vs col O da planilha.")

    print("\n=== VALIDAÇÃO DE PONTOS ===")
    for v in res.validacao:
        status_str = "OK" if v.ok else ("SEM ESPERADO" if v.esperado is None else "DIVERGE")
        print(f"  {v.nome:12s} | calc {v.calculado:>4d} | esperado {str(v.esperado):>4s} | {status_str}")

    if res.todos_ok:
        print("\nImportação concluída com sucesso — todos os totais conferem.")
    else:
        print("\nERRO: totais divergem do esperado.")
        sys.exit(1)


if __name__ == "__main__":
    main()
