"""Backfill de palpites do mata-mata a partir da planilha.

Por que existe (e por que NÃO reusa ``importar_planilha.py``):
    - O importador da fase de grupos é fixo nas linhas 2–73 e nas 3 rodadas de
      grupo; ignora o mata-mata (linha 74+).
    - Ele também PROTEGE jogos já encerrados (não reescreve os palpites deles) —
      mas no mata-mata é exatamente isso que precisamos: a tabela de palpites foi
      feita antes dos jogos e subiu atrasada, com a rodada já fechada e alguns
      jogos já realizados.
    - O caminho normal (``salvar_palpite``) revalida o PRAZO e barra a gravação
      numa rodada fechada.

O que este script faz:
    Lê as linhas do mata-mata da aba OFICIAL (a partir da linha 74) e, para cada
    aba de jogador, grava os palpites direto na tabela ``palpites`` — contornando
    o prazo de propósito (são palpites travados antes do jogo, apenas subidos
    atrasado). O casamento jogo↔linha é por NOME dos times dentro das rodadas de
    mata-mata (``Rodada.ordem >= 4``).

Cálculo dos pontos:
    - Jogo já ENCERRADO no banco → calcula ``pontos`` agora, pela LEGENDA, a
      partir do resultado oficial que está no banco (autoritativo — veio da
      ESPN/admin).
    - Jogo ainda não encerrado → grava ``pontos = 0``. Quando a ESPN lançar o
      resultado, ``admin.lancar_resultado`` recalcula TODOS os palpites do jogo
      (inclusive estes) — então a pontuação fica correta independente da ordem.

NUNCA toca no placar/status do jogo nem na configuração da rodada. Idempotente:
re-rodar faz upsert por (usuário, jogo) e recalcula os pontos pelo resultado atual.

Uso:
    # dry-run (padrão): mostra o que faria, sem gravar nada
    python scripts/importar_palpites_mata_mata.py

    # grava de fato
    python scripts/importar_palpites_mata_mata.py --commit

    # outra planilha (ex.: quando sair a das oitavas)
    python scripts/importar_palpites_mata_mata.py --commit --arquivo "NOME.xlsx"

Requisitos:
    pip install openpyxl   (já é dependência do importador da fase de grupos)
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable

# ---------------------------------------------------------------------------
# Bootstrap de sys.path para permitir ``import app...`` ao rodar diretamente.
# ---------------------------------------------------------------------------
_PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_PROJECT_ROOT))

# Console do Windows quebra com acentos/emoji em cp1252; força UTF-8 quando dá.
for _fluxo in (sys.stdout, sys.stderr):
    try:
        _fluxo.reconfigure(encoding="utf-8")  # type: ignore[union-attr]
    except (AttributeError, ValueError):
        pass

import openpyxl  # noqa: E402
from openpyxl.worksheet.worksheet import Worksheet  # noqa: E402
from sqlalchemy import select  # noqa: E402
from sqlalchemy.orm import Session  # noqa: E402

from app.database import SessionLocal  # noqa: E402
from app.models import Jogo, Palpite, Rodada, Usuario  # noqa: E402
from app.services.dashboard import STATUS_ENCERRADO  # noqa: E402
from app.services.scoring import calcular_pontos  # noqa: E402
from scripts.importar_planilha import JOGADORES, XLSX_PATH  # noqa: E402

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

ABA_OFICIAL: str = "OFICIAL"

# Primeira linha do mata-mata na aba OFICIAL (a fase de grupos vai até a 73).
MATA_MATA_LINHA_INICIO: int = 74

# Rodadas com ordem >= esta são mata-mata (ver app/services/espn.py).
ORDEM_PRIMEIRO_MATA_MATA: int = 4

# Planilha padrão: a de "ATÉ A FINAL 16 AVOS" na pasta import/.
XLSX_PADRAO: Path = _PROJECT_ROOT / "import" / "COPA PHIBRA 2026 OFICIAL ATÉ A FINAL 16 AVOS.xlsx"


# ---------------------------------------------------------------------------
# Resultado
# ---------------------------------------------------------------------------


@dataclass(slots=True)
class ResultadoBackfill:
    """Resumo do backfill — o caller (CLI) decide o que mostrar."""

    arquivo: str
    commit: bool
    palpites_criados: int = 0
    palpites_atualizados: int = 0
    jogos_com_palpite: int = 0
    jogos_encerrados_pontuados: int = 0
    # Por jogador: (nome, qtd_palpites_no_mata_mata)
    por_jogador: list[tuple[str, int]] = field(default_factory=list)
    # Linhas com palpite mas sem Jogo correspondente no banco (warning).
    sem_jogo: list[tuple[str, int, str, str]] = field(default_factory=list)
    # Linhas de detalhe p/ log: (jogador, casa, visitante, pc, pv, status, pontos)
    detalhes: list[tuple[str, str, str, int, int, str, int]] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Leitura da planilha
# ---------------------------------------------------------------------------


def _ler_linhas_oficial_mata_mata(ws: Worksheet) -> dict[int, tuple[str, str]]:
    """row -> (time_casa, time_visitante) das linhas de mata-mata da aba OFICIAL.

    Inclui linhas com placeholders ("Venc. ...") — elas simplesmente não casam
    com nenhum Jogo do banco e não têm palpite, então são ignoradas adiante.
    """
    linhas: dict[int, tuple[str, str]] = {}
    for r in range(MATA_MATA_LINHA_INICIO, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=r, max_row=r, min_col=1, max_col=7, values_only=True))[0]
        time_casa = str(row[2]).strip() if row[2] is not None else ""
        time_visitante = str(row[6]).strip() if row[6] is not None else ""
        if not time_casa or not time_visitante:
            continue
        linhas[r] = (time_casa, time_visitante)
    return linhas


def _ler_palpites_jogador(
    ws: Worksheet,
    nome_aba: str,
    linhas_oficial: dict[int, tuple[str, str]],
) -> dict[int, tuple[int, int]]:
    """row -> (gols_casa, gols_visitante) dos palpites de um jogador no mata-mata.

    Só inclui linhas onde D e F estão preenchidos. Valida alinhamento C/G contra
    a aba OFICIAL e aborta (RuntimeError) em qualquer divergência — nunca importa
    palpite desalinhado.
    """
    palpites: dict[int, tuple[int, int]] = {}
    for r, (tc_oficial, tv_oficial) in linhas_oficial.items():
        row = list(ws.iter_rows(min_row=r, max_row=r, min_col=1, max_col=7, values_only=True))[0]
        tc = str(row[2]).strip() if row[2] is not None else ""
        tv = str(row[6]).strip() if row[6] is not None else ""
        if tc != tc_oficial or tv != tv_oficial:
            raise RuntimeError(
                f"Desalinhamento na aba {nome_aba!r} linha {r}: "
                f"esperado {tc_oficial!r} x {tv_oficial!r}, encontrado {tc!r} x {tv!r}. "
                "Abortar — não importe palpites desalinhados."
            )
        d_raw, f_raw = row[3], row[5]
        if d_raw is None or f_raw is None:
            continue  # jogador não palpitou este jogo
        palpites[r] = (int(d_raw), int(f_raw))
    return palpites


# ---------------------------------------------------------------------------
# Backfill
# ---------------------------------------------------------------------------


def importar_palpites_mata_mata(
    commit: bool = False,
    xlsx_path: Path | None = None,
    session_factory: Callable[[], Session] | None = None,
) -> ResultadoBackfill:
    """Grava os palpites do mata-mata da planilha no banco.

    Args:
        commit: se False (padrão), faz dry-run — calcula tudo mas dá rollback.
        xlsx_path: caminho do .xlsx. Padrão: XLSX_PADRAO (16 avos).
        session_factory: fábrica de sessão SQLAlchemy. Padrão: SessionLocal.
    """
    _xlsx = xlsx_path if xlsx_path is not None else XLSX_PADRAO
    _session_factory = session_factory if session_factory is not None else SessionLocal

    if not _xlsx.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {_xlsx}")

    wb = openpyxl.load_workbook(str(_xlsx), data_only=True)
    for aba in [ABA_OFICIAL] + list(JOGADORES.keys()):
        if aba not in wb.sheetnames:
            raise RuntimeError(f"Aba {aba!r} não encontrada na planilha.")

    linhas_oficial = _ler_linhas_oficial_mata_mata(wb[ABA_OFICIAL])

    res = ResultadoBackfill(arquivo=_xlsx.name, commit=commit)
    jogos_com_palpite: set[int] = set()
    jogos_encerrados_pontuados: set[int] = set()

    db = _session_factory()
    try:
        # Índice (time_casa, time_visitante) -> Jogo, restrito ao mata-mata.
        jogos_ko = db.execute(
            select(Jogo)
            .join(Rodada, Jogo.rodada_id == Rodada.id)
            .where(Rodada.ordem >= ORDEM_PRIMEIRO_MATA_MATA)
        ).scalars().all()
        jogo_por_times: dict[tuple[str, str], Jogo] = {
            (j.time_casa, j.time_visitante): j for j in jogos_ko
        }

        # Usuários por aba (username = nome de exibição em minúsculas).
        usuario_por_aba: dict[str, Usuario] = {}
        for aba_nome, nome_display in JOGADORES.items():
            usuario = db.scalar(select(Usuario).where(Usuario.username == nome_display.lower()))
            if usuario is None:
                raise RuntimeError(
                    f"Usuário {nome_display!r} (username={nome_display.lower()!r}) não existe. "
                    "Rode antes o importador da fase de grupos para criar os jogadores."
                )
            usuario_por_aba[aba_nome] = usuario

        for aba_nome in JOGADORES:
            usuario = usuario_por_aba[aba_nome]
            palpites_aba = _ler_palpites_jogador(wb[aba_nome], aba_nome, linhas_oficial)
            qtd_jogador = 0

            for row_num, (pc, pv) in palpites_aba.items():
                tc, tv = linhas_oficial[row_num]
                jogo = jogo_por_times.get((tc, tv))
                if jogo is None:
                    # Palpite numa linha sem jogo correspondente no banco (ex.:
                    # confronto ainda não definido). Registra e segue.
                    res.sem_jogo.append((aba_nome, row_num, tc, tv))
                    continue

                # Pontos: do resultado atual se encerrado; senão 0 (a ESPN
                # recalcula quando lançar o resultado).
                if (
                    jogo.status == STATUS_ENCERRADO
                    and jogo.gols_casa is not None
                    and jogo.gols_visitante is not None
                ):
                    pontos = calcular_pontos(pc, pv, jogo.gols_casa, jogo.gols_visitante)
                    jogos_encerrados_pontuados.add(jogo.id)
                else:
                    pontos = 0

                palpite = db.scalar(
                    select(Palpite).where(
                        Palpite.usuario_id == usuario.id,
                        Palpite.jogo_id == jogo.id,
                    )
                )
                if palpite is None:
                    db.add(
                        Palpite(
                            usuario_id=usuario.id,
                            jogo_id=jogo.id,
                            gols_casa=pc,
                            gols_visitante=pv,
                            pontos=pontos,
                        )
                    )
                    res.palpites_criados += 1
                else:
                    palpite.gols_casa = pc
                    palpite.gols_visitante = pv
                    palpite.pontos = pontos
                    res.palpites_atualizados += 1

                jogos_com_palpite.add(jogo.id)
                qtd_jogador += 1
                res.detalhes.append((aba_nome, tc, tv, pc, pv, jogo.status, pontos))

            res.por_jogador.append((JOGADORES[aba_nome], qtd_jogador))

        res.jogos_com_palpite = len(jogos_com_palpite)
        res.jogos_encerrados_pontuados = len(jogos_encerrados_pontuados)

        if commit:
            db.commit()
        else:
            db.rollback()
        return res
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--commit",
        action="store_true",
        help="Grava de fato. Sem esta flag, roda em dry-run (nada é gravado).",
    )
    parser.add_argument(
        "--arquivo",
        default=None,
        help="Nome do .xlsx na pasta import/ (padrão: a planilha de 16 avos).",
    )
    args = parser.parse_args()

    xlsx_path = (_PROJECT_ROOT / "import" / args.arquivo) if args.arquivo else XLSX_PADRAO

    modo = "COMMIT (gravando)" if args.commit else "DRY-RUN (nada gravado)"
    print(f"=== Backfill palpites mata-mata — {modo} ===")
    print(f"Planilha: {xlsx_path.name}\n")

    res = importar_palpites_mata_mata(commit=args.commit, xlsx_path=xlsx_path)

    print("Palpites por jogador (mata-mata):")
    for nome, qtd in res.por_jogador:
        print(f"  {nome:12s} | {qtd:>2d} palpites")

    print(
        f"\nResumo: {res.palpites_criados} criados, {res.palpites_atualizados} atualizados | "
        f"{res.jogos_com_palpite} jogos com palpite | "
        f"{res.jogos_encerrados_pontuados} jogos já encerrados (pontuados agora)."
    )

    if res.sem_jogo:
        print(f"\nAVISO — {len(res.sem_jogo)} palpite(s) sem jogo correspondente no banco:")
        for aba, row, tc, tv in res.sem_jogo:
            print(f"  {aba:12s} L{row}: {tc} x {tv}")

    if not args.commit:
        print("\n(DRY-RUN: nada foi gravado. Rode com --commit para persistir.)")


if __name__ == "__main__":
    main()
